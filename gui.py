from pathlib import Path
from tkinter import filedialog, messagebox
import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage


OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / "assets" / "frame0"
ICON_PATH = ASSETS_PATH / "icons8-create-100.ico"


def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)


# Function to handle button1 click event
def select_folder():
    global input_folder_path
    input_folder_path = filedialog.askdirectory() # Open folder selection dialog
    if input_folder_path:  # Check if a folder is selected
        print("Selected folder:", input_folder_path)
        #process_csv_files(folder_path)
    

# Function to handle button2 click event
def select_word_document():
    word_file_path = filedialog.askopenfilename(
        filetypes=[("Word Documents", "*.docx;*.doc")])  # Open file selection dialog for Word documents
    if word_file_path:  # Check if a file is selected
        print("Selected Word document:", word_file_path)
        # Here you can save the file_path to use later
        # Display a warning message
        messagebox.showwarning("Uyarı", "Lütfen, en son revizyon IRS dokümanını seçtiğinizden emin olunuz!")

# Function to handle button3 click event
def select_output_folder():
    global output_folder_path
    output_folder_path = filedialog.askdirectory()  # Open folder selection dialog
    
    if output_folder_path:  # Check if a folder is selected
        print("Selected output folder:", output_folder_path)
        global output_path_excel
        output_path_excel = Path(output_folder_path) / "merged_data.xlsx"
        print("Excel will be printed:", output_path_excel)
        # Assign the folder_path to a variable for later use
        # Example: output_folder_path = folder_path


def call_functions():
    process_csv_files() and apply_inverse_transformation()
    


def process_csv_files():

    # İşlem yapılacak klasör yolu
    #input_folder = "C:/Users/alika/Desktop/Reporter_Project/Inputs"

    # Tüm .csv dosyalarını al
    csv_files = glob.glob(input_folder_path + "/*.csv")

    # İlk .csv dosyasını oku ve 'Min' ve 'Max' sütunlarını ekle
    merged_df = pd.read_csv(csv_files[0])
    merged_df.insert(loc=7, column='Min', value='')
    merged_df.insert(loc=8, column='Max', value='')

    # 'Dev' ve 'Out' sütunlarını sil
    merged_df = merged_df.drop(columns=['Dev' , 'Actual' , 'Out' , 'Check'])

    # Diğer .csv dosyalarını oku ve 'Check' sütunlarını ekleyin
    for file in csv_files[0:]:
        # Dosya adını al ve 'Check' sütununu oku
        column_name = file.split("/")[-1].split(".csv")[0][-1]  # Dosya adının son karakteri
        check_column = pd.read_csv(file)['Check']
        # DataFrame'e 'Check' sütununu ekle
        merged_df[column_name] = check_column

    # 'Min' ve 'Max' sütunlarındaki boş hücrelere minimum ve maksimum değerleri yaz
    for index, row in merged_df.iterrows():
        min_value = row[9:].min()  # İlk 10 sütunun dışındaki sütunlar 'Check' sütunlarıdır
        max_value = row[9:].max()
        merged_df.at[index, 'Min'] = min_value
        merged_df.at[index, 'Max'] = max_value

    # Kolon adlarını güncelle
    columns_mapping = {col: col.split(" ")[0] for col in merged_df.columns[9:]}
    merged_df = merged_df.rename(columns=columns_mapping)

    # Birleştirilmiş veriyi göster
    print(merged_df)

    # Birleştirilmiş veriyi Excel dosyasına dönüştür ve kaydet
    #global output_path_excel
    #output_path_excel = output_folder_path / "merged_data.xlsx"
    merged_df.to_excel(output_path_excel, index=False, engine='openpyxl', float_format="%.2f", header=True)

    # Excel dosyasını aç ve hücre tiplerini belirle
    wb = load_workbook(output_path_excel)
    ws = wb.active

    # Mavi ve kırmızı renkler
    blue_fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Her satır için koşulları kontrol et ve hücreleri uygun şekilde boyayın
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for index, cell in enumerate(row[2:], start=2):  # 'Min' kolonundan başlayarak
            nominal = float(row[2].value)  # Nominal değeri al
            tol_minus = float(row[3].value)  # Tol- değeri al
            tol_plus = float(row[4].value)  # Tol+ değeri al
            cell_value = row[index].value  # Hücrenin değeri
            try: #çalışmıyor
                cell_value = float(cell_value)
                if cell_value < (nominal - tol_minus):  # Koşul 1: Nominal - Tol-
                    cell.fill = blue_fill  # Maviye boyama
                elif cell_value > (nominal + tol_plus):  # Koşul 2: Nominal + Tol+
                    cell.fill = red_fill  # Kırmızıya boyama
            except ValueError:
                pass

    # Excel dosyasını kaydet
    wb.save(output_path_excel)
#process_csv_files()


def apply_inverse_transformation():
    # Excel dosyasını oku
    input_excel = Path(output_folder_path) /"merged_data.xlsx"
    output_excel = Path(output_folder_path) /"merged_data.xlsx"
    df = pd.read_excel(input_excel)

    # İstenmeyen sütunları sil
    df = df.drop(columns=['Property', 'Nominal', 'Tol -', 'Tol +', 'Min', 'Max'])
    print(df)
    # Transpozunu al
    df = df.T
    print("###############################################")
    print(df)

    # Yeni sütun ekle ve ilk satırı 'Parca No' olarak ayarla
    num_rows = len(df)
    #df.insert(0, 'Parca No', range(1, num_rows + 1))

    # Devrik dönüşüm uygula
    #df_inverse = df.apply(lambda x: 1/x if x.dtype == 'float' else x)
    

    # Yeni Excel dosyasını oluştur ve verileri kaydet
    with pd.ExcelWriter(output_excel, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='TO_WORD', index=False)
     
     # Silme işlemi
        # Excel dosyasını aç
        wb = writer.book
        ws = wb['TO_WORD']
        # İlk satırı sil
        ws.delete_rows(1)

# Yeni sütun ekle ve ilk satırı 'Parca No' olarak ayarla
        ws.insert_cols(1)
        ws.cell(row=1, column=1, value='Parca No')
        # Alt satırlara doğru 'Parca No' değerlerini ekle
        for i in range(2, num_rows + 1):
            ws.cell(row=i, column=1, value=i-1)

# Kullanım örneği
#input_excel = output_folder_path /"merged_data.xlsx"
#output_excel = output_folder_path /"merged_data.xlsx"  # Aynı dosya üzerine kaydedilecek
#apply_inverse_transformation(input_excel, output_excel)


window = Tk()
window.title("Atos Rapor Birleştirici")
window.iconbitmap(ICON_PATH)  # Set the icon for the window

window.geometry("1288x625")
window.configure(bg = "#FFFFFF")


canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 625,
    width = 1288,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)
canvas.create_rectangle(
    0.0,
    0.0,
    632.0,
    792.0,
    fill="#1E275C",
    outline="")

canvas.create_rectangle(
    634.0,
    0.0,
    1288.0,
    792.0,
    fill="#A7A9AC",
    outline="")

canvas.create_text(
    11.0,
    61.0,
    anchor="nw",
    text="Atos Rapor Birleştirici' ye \nHoş Geldiniz!",
    fill="#FFFFFF",
    font=("CourierPrime Regular", 32 * -1)
)

canvas.create_text(
    11.0,
    190.0,
    anchor="nw",
    text="Atos rapor birleştiricisi\nbirden fazla parçaya ait ölçümleri\ntek bir veri setine dönüştürür ve\nIRS doldurmak için kullanır.",
    fill="#FFFFFF",
    font=("CourierPrime Regular", 32 * -1)
)

canvas.create_text(
    769.0,
    61.0,
    anchor="nw",
    text="Detayları belirtiniz.",
    fill="#FFFFFF",
    font=("CourierPrime Regular", 32 * -1)
)

canvas.create_rectangle(
    749.0,
    126.0,
    1204.0,
    216.0,
    fill="#CAC9C9",
    outline="")

canvas.create_rectangle(
    749.0,
    256.0,
    1204.0,
    346.0,
    fill="#CAC9C9",
    outline="")

canvas.create_rectangle(
    749.0,
    386.0,
    1204.0,
    476.0,
    fill="#CAC9C9",
    outline="")

canvas.create_text(
    769.0,
    122.0,
    anchor="nw",
    text="Atos Rapor Dosyaları",
    fill="#FFFFFF",
    font=("CourierPrime Regular", 32 * -1)
)

canvas.create_text(
    769.0,
    252.0,
    anchor="nw",
    text="IRS Dosyaları",
    fill="#FFFFFF",
    font=("CourierPrime Regular", 32 * -1)
)

canvas.create_text(
    769.0,
    386.0,
    anchor="nw",
    text="Çıktı Yolu",
    fill="#FFFFFF",
    font=("CourierPrime Regular", 32 * -1)
)

button_image_1 = PhotoImage(
    file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=select_folder,   # Assign the function
    relief="flat"

)
button_1.place(
    x=1109.0,
    y=158.0,
    width=73.0,
    height=58.0
)

button_image_2 = PhotoImage(
    file=relative_to_assets("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=call_functions,
    relief="flat"
    
)
button_2.place(
    x=927.0,
    y=518.0,
    width=80.0,
    height=80.0
)

button_image_3 = PhotoImage(
    file=relative_to_assets("button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    command=select_output_folder,  # Assign the function
    relief="flat"
)
button_3.place(
    x=1109.0,
    y=416.0,
    width=68.0,
    height=60.0
)

button_image_4 = PhotoImage(
    file=relative_to_assets("button_4.png"))
button_4 = Button(
    image=button_image_4,
    borderwidth=0,
    highlightthickness=0,
    command=select_word_document,  # Assign the function
    relief="flat"
)
button_4.place(
    x=1109.0,
    y=286.0,
    width=68.0,
    height=60.0
)
window.resizable(False, False)
window.mainloop()
